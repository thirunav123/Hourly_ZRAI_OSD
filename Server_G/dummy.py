from importlib.metadata import requires
import os,socket,queue,datetime,time,threading,requests,json,schedule
from openpyxl import Workbook,load_workbook
from smtplib import SMTP
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

setting_file=open('server_settings.txt','r')
print("MT: I am a Server")
filedic={}
for line in setting_file:
    file_data=line.strip().split('===')
    a=file_data[0]
    b=file_data[1]
    filedic[a]=b
setting_file.close()
ipaddress_of_system=filedic.pop('ipaddress_of_server_system')
port_to_listen=int(filedic.pop('port_to_listen'))
shiftA_start=filedic.pop('shiftA_start_time')
shiftB_start=filedic.pop('shiftB_start_time')
shiftC_start=filedic.pop('shiftC_start_time')
filename_of_i_excel=filedic.pop('filename_of_idle_time_excel_sheet')
filename_of_h_excel=filedic.pop('filename_of_hourly_excel_sheet')
mfg_webhook_url1=filedic.pop('mfg_webhook_url_level1')
mfg_webhook_url2=filedic.pop('mfg_webhook_url_level2')
mfg_webhook_url3=filedic.pop('mfg_webhook_url_level3')
bytes_to_split=int(filedic.pop('bytes_to_split'))
level1_name=filedic.pop('level1_name')
level2_name=filedic.pop('level2_name')
level3_name=filedic.pop('level3_name')
hourly_sheet_name=filedic.pop('hourly_sheet_name')
header_row_h=int(filedic.pop('header_row_hourly'))
total_columns=int(filedic.pop('total_columns'))
mfg_space_url=filedic.pop('mfg_space_url')
scheduler_delay=int(filedic.pop("scheduler_delay_in_milliseconds"))/1000
error_wait=int(filedic.pop("error_wait_in_milliseconds"))/1000
h_chat_queue_timeout=int(filedic.pop("hourly_chat_queue_timeout_in_seconds"))
mail_trig_delay=int(filedic.pop("mail_trigger_delay_after_next_shift_start_in_seconds"))
common_letters=int(filedic.pop("common_letters_count_from_begin_in_pq_and_it_columns"))
print("MT: Settings file read")

total_file= open("total.txt", "r")
mail_file = open("mail.txt", "r")
summary_a_file= open("summary_A.txt", "r")
summary_b_file= open("summary_B.txt", "r")
summary_c_file= open("summary_C.txt", "r")

total_dic={}
list_of_mails= []
summary_a_list= []
summary_b_list= []
summary_c_list= []
for line in total_file:
    file_data=line.strip().split('===')
    a=int(file_data[0])
    # b=.strip().split(",")
    b=map(int,file_data[1].strip().split(","))
    total_dic[a]=b
for line in mail_file:
  list_of_mails.append(line.strip())
for line in summary_a_file:
  summary_a_list.append(line.strip())
for line in summary_b_file:
  summary_b_list.append(line.strip())
for line in summary_c_file:
  summary_c_list.append(line.strip())

total_file.close()
mail_file.close()
summary_a_file.close()
summary_b_file.close()
summary_c_file.close()
summary_dic={"A":summary_a_list,"B":summary_b_list,"C":summary_c_list}
print("MT: Total, Mail, Summary_A, Summary_B, Summary_C file read")
A=list(map(int,shiftA_start.strip().split(":")))
B=list(map(int,shiftB_start.strip().split(":")))
C=list(map(int,shiftC_start.strip().split(":")))
data_queue=queue.Queue()
it_excel_queue=queue.Queue()
it_chat_queue=queue.Queue()
h_chat_queue=queue.Queue()
h_excel_queue=queue.Queue()
error_register_queue=queue.Queue()
h_excel_queue.put([str(104),str(17),9,40,2,round(time.time())])

def move_excel_h():
    # line_column,date_column,shift_column,time_h_column,time_i_column,on_off_c_column,reqiured_row=1,1,1,1,1,1,1
    reqiured_row=1
    write_flag=False
    while True:
        # try:
            transfer_dic={}
            plc,h,on_off_c,pq,it,time_eve=h_excel_queue.get()
            t=datetime.datetime.fromtimestamp(time_eve)
            current_time=datetime.time(t.hour,t.minute,t.second)
            if current_time<datetime.time(A[0],A[1],A[2]):
                dp=1
            else:
                dp=0
            # plc,"Total",pq,it,time_of_eve
            date=t-datetime.timedelta(days=dp)
            date=date.strftime("%d-%m-%Y")
            # time_now=now.strftime("%I.%M.%S_%p")
            transfer_dic["LINE"]=filedic[plc]
            transfer_dic["DATE"]=date
            transfer_dic["HOUR"]=filedic[h]
            transfer_dic["SHIFT"]="A"
            transfer_dic["ON/OFF_Count"]=on_off_c
            transfer_dic["PQ"]=pq
            # line_pq=filedic [plc]+"_pq"
            # line_it=filedic[plc]+"_it"
            # time_h=h+"_pq"
            # time_i=h+"_it"
            # transfer_dic[time_h]=pq
            transfer_dic["IT"]=round(it/60,2)
            
            # if not os.path.isfile(filename_of_excel):
            if not os.path.isfile(filename_of_h_excel):
                wb=Workbook()
            else:
                wb=load_workbook(filename_of_h_excel)
            station=hourly_sheet_name
            # if station_list!=wb.sheetnames:
            #     for station in station_list:
            if not station in wb.sheetnames:
                wb.create_sheet(station)
                ws=wb[station]
                for i in range(1,total_columns+1):
                    ws.cell(row=header_row_h,column=i).value=filedic[str(i)]
            ws=wb[station]        
            xl_headers=[]
            for i in ws[header_row_h]:
                xl_headers.append(i.value)
            # mc=ws.max_column
            # mr=ws.max_row
            # for i in transfer_dic:
            #     if i not in xl_headers:
            #         ws.cell(1,mc+1).value=i
            #         xl_headers.append(i)
            #         mc=ws.max_column
            date_column=int(filedic["date_column"])
            line_column=int(filedic["line_column"])
            # for i in range(1,mc+ 1):
            #     # if ws.cell(row=header_row_h,coulmn=i):
            #     v=ws.cell(row=header_row_h,column=i).value
            #     if v=="LINE":
            #         line_column=i
            #         # print(i)
            #     if v=="DATE":
            #         date_column=i
                    # print(i)
                # if v=="SHIFT":
                #     shift_column=i
                    # print(i)
                # if v==time_h:
                #     time_h_column=i
                #     # print(i)
                # if v==time_i:
                #     time_i_column=i
                # if v=="ON/OFF_Count":
                #     on_off_c_column=i
                    # print(i)
            compare_list=[transfer_dic["LINE"],transfer_dic["DATE"]]
            value_list=[]
            exist_flag=False
            h=int(h)
            for row in ws.rows:
                for cell in row:
                    value_list.append(cell.value)
                # print(type(row))
                check = all(item in value_list for item in compare_list)
                if check:
                    print("Exist")
                    exist_flag=True
                    reqiured_row=cell.row
                    ws.cell(row=reqiured_row,column=h).value=transfer_dic["PQ"]
                    ws.cell(row=reqiured_row,column=h+1).value=transfer_dic["IT"]
                    if transfer_dic["SHIFT"]=="A":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countA"])).value=transfer_dic["ON/OFF_Count"]
                    elif transfer_dic["SHIFT"]=="B":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countB"])).value=transfer_dic["ON/OFF_Count"]
                    else:
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countC"])).value=transfer_dic["ON/OFF_Count"]
                    break
            mr=ws.max_row
            # print( line_column,date_column,shift_column,time_h_column,time_i_column)

            if not exist_flag:
                print("Not Exist")
                reqiured_row=mr+1
                ws.cell(row=reqiured_row,column=line_column).value=transfer_dic["LINE"]
                ws.cell(row=reqiured_row,column=date_column).value=transfer_dic["DATE"]
                ws.cell(row=reqiured_row,column=h).value=transfer_dic["PQ"]
                ws.cell(row=reqiured_row,column=h+1).value=transfer_dic["IT"]
                if transfer_dic["SHIFT"]=="A":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countA"])).value=transfer_dic["ON/OFF_Count"]
                elif transfer_dic["SHIFT"]=="B":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countB"])).value=transfer_dic["ON/OFF_Count"]
                else:
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countC"])).value=transfer_dic["ON/OFF_Count"]
            raw_list=list(ws.rows)
            for t in total_dic:
                total=0
                for a in total_dic[t]:
                    v=raw_list[reqiured_row-1][a-1].value
                    if not v is None:
                        total=total+int(v)
                ws.cell(row=reqiured_row,column=t).value=total
                # if ws.cell(row=i,column=line_column).value==transfer_dic["LINE"]:
                    # if ws.cell(row=i,column=line_column).value==transfer_dic["LINE"]:
                
            # it_chat_queue.put([2,transfer_dic["LINE"],transfer_dic["DATE"],transfer_dic["SHIFT"],transfer_dic[time_h],transfer_dic[time_i]])

            # for i in xl_headers:.lo0  
            # .0   
            #     for j in transfer_dic:
            #         if i==j:
            #             ws.cell(mr+1,xl_headers.index(i)+1).value=excel_dic[i]
            while True:
                try:
                    wb.save(filename_of_h_excel)
                    print(f"EXH: Hourly and Idle time data saved successfully {time.strftime('%d-%m-%Y_%I.%M.%S_%p')} {transfer_dic['LINE']}")
                except:
                    # wb.save(filename_of_excel)
                    print(f"EXH: Data not saved, Close the excel file({filename_of_h_excel}) if it is opened. Retrying to save...")
                    time.sleep(3)
                    continue
                break
            wb.close()
            write_flag=False
        # except Exception as e:
        #     if not write_flag:
        #         error_register_queue.put(f'EXH: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
        #         write_flag=True
        #         print(f'EXH: Error acquired and recorded. Error: {e}')
        #     else: 
        #         print(f'EXH: Error acquired and already recorded. Error: {e}')
        #     time.sleep(error_wait)

move_ex_h = threading.Thread(target=move_excel_h,daemon=True)
move_ex_h.start()
print("MT: Move_excel_h thread started")
while True:
  time.sleep(1)