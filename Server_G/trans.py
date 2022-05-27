from importlib.metadata import requires
import os,socket,queue,datetime,time,threading,requests,json,schedule
from openpyxl import Workbook,load_workbook
from smtplib import SMTP
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

setting_file=open('server_settings.txt','r')
print("MT: I am a Server")
filedic={}
for line in setting_file:
    file_data=line.strip().split('===')
    a=file_data[0]
    b=file_data[1]
    filedic[a]=b
setting_file.close()
ipaddress_of_system=filedic.pop('ipaddress_of_server_system')
port_to_listen=int(filedic.pop('port_to_listen'))
shiftA_start=filedic.pop('shiftA_start_time')
shiftB_start=filedic.pop('shiftB_start_time')
shiftC_start=filedic.pop('shiftC_start_time')
filename_of_i_excel=filedic.pop('filename_of_idle_time_excel_sheet')
filename_of_h_excel=filedic.pop('filename_of_hourly_excel_sheet')
mfg_webhook_url1=filedic.pop('mfg_webhook_url_level1')
mfg_webhook_url2=filedic.pop('mfg_webhook_url_level2')
mfg_webhook_url3=filedic.pop('mfg_webhook_url_level3')
bytes_to_split=int(filedic.pop('bytes_to_split'))
level1_name=filedic.pop('level1_name')
level2_name=filedic.pop('level2_name')
level3_name=filedic.pop('level3_name')
hourly_sheet_name=filedic.pop('hourly_sheet_name')
header_row_h=int(filedic.pop('header_row_hourly'))
total_columns=int(filedic.pop('total_columns'))
mfg_space_url=filedic.pop('mfg_space_url')
scheduler_delay=int(filedic.pop("scheduler_delay_in_milliseconds"))/1000
error_wait=int(filedic.pop("error_wait_in_milliseconds"))/1000
h_chat_queue_timeout=int(filedic.pop("hourly_chat_queue_timeout_in_seconds"))
mail_trig_delay=int(filedic.pop("mail_trigger_delay_after_next_shift_start_in_seconds"))
common_letters=int(filedic.pop("common_letters_count_from_begin_in_pq_and_it_columns"))
print("MT: Settings file read")

total_file= open("total.txt", "r")
mail_file = open("mail.txt", "r")
summary_a_file= open("summary_A.txt", "r")
summary_b_file= open("summary_B.txt", "r")
summary_c_file= open("summary_C.txt", "r")

total_dic={}
list_of_mails= []
summary_a_list= []
summary_b_list= []
summary_c_list= []
for line in total_file:
    file_data=line.strip().split('===')
    a=int(file_data[0])
    # b=.strip().split(",")
    b=map(int,file_data[1].strip().split(","))
    total_dic[a]=b
for line in mail_file:
  list_of_mails.append(line.strip())
for line in summary_a_file:
  summary_a_list.append(line.strip())
for line in summary_b_file:
  summary_b_list.append(line.strip())
for line in summary_c_file:
  summary_c_list.append(line.strip())

total_file.close()
mail_file.close()
summary_a_file.close()
summary_b_file.close()
summary_c_file.close()
summary_dic={"A":summary_a_list,"B":summary_b_list,"C":summary_c_list}
print("MT: Total, Mail, Summary_A, Summary_B, Summary_C file read")
A=list(map(int,shiftA_start.strip().split(":")))
B=list(map(int,shiftB_start.strip().split(":")))
C=list(map(int,shiftC_start.strip().split(":")))
data_queue=queue.Queue()
it_excel_queue=queue.Queue()
it_chat_queue=queue.Queue()
h_chat_queue=queue.Queue()
h_excel_queue=queue.Queue()
error_register_queue=queue.Queue()


def get_shift(ct):
    startA=datetime.time(A[0],A[1],A[2])
    startB=datetime.time(B[0],B[1],B[2])
    startC=datetime.time(C[0],C[1],C[2])
    if startA<ct<startB:
        return 'A'
    if startB<ct<startC:
        return 'B'
    else:
        return 'C'

def error_register():
    while True:
        err=error_register_queue.get()
        with open('readme.txt', 'a') as f:
            f.write(err)
        f.close()

def data_decode():
    write_flag=False
    while True:
        try:
            data=data_queue.get()
            # data=bytearray(data)
            if data[0]==1:
                    plc=str(data[1])
                    status=data[2]
                    initial_time,rt = [int.from_bytes(data[i:i + bytes_to_split],"big") for i in range(3, len(data), bytes_to_split)]
                    # print(data_list)
                    # initial_time,rt,status=data_list
                    from_time_format=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                    idle_time=round(rt-initial_time)/60
                    recv_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(rt))
                    if status==0:
                        it_excel_queue.put(["Instance",plc,from_time_format,recv_time,idle_time])
                        print("DD: i_Instance ",filedic[plc])
                    elif status==1:
                        it_excel_queue.put([1,plc,from_time_format,recv_time,idle_time])
                        it_chat_queue.put([[1,"O"],plc,from_time_format,recv_time,idle_time])
                        print("DD: i_LLLLLLLLLLLLLLLLLLLLLevel1 ",filedic[plc])
                    elif status==2:
                        it_excel_queue.put([2,plc,from_time_format,recv_time,idle_time])
                        it_chat_queue.put([[2,"O"],plc,from_time_format,recv_time,idle_time])
                        print("DD: i_LLLLLLLLLLLLLLLLLLLLLevel2 ",filedic[plc])
                    elif status==3:
                        it_excel_queue.put([3,plc,from_time_format,recv_time,idle_time])
                        it_chat_queue.put([[3,"O"],plc,from_time_format,recv_time,idle_time])
                        print("DD: i_LLLLLLLLLLLLLLLLLLLLLevel3 ",filedic[plc])
                    else:
                        it_excel_queue.put(["Cleared",plc,from_time_format,recv_time,idle_time])
                        if status==7:
                            it_chat_queue.put([[3,"C"],plc,from_time_format,recv_time,idle_time])
                        elif status==6:
                            it_chat_queue.put([[2,"C"],plc,from_time_format,recv_time,idle_time])
                        elif status==5:
                            it_chat_queue.put([[1,"C"],plc,from_time_format,recv_time,idle_time])
                        elif status==4:
                            print("DD: Instance Cleared")
                        print("DD: Idle time cleared ",filedic[plc])
            elif data[0]==2:
                plc=str(data[1])
                h=str(data[2])
                on_off_c=str(data[3])
                pq,it,tpq,tit,time_of_eve = [int.from_bytes(data[i:i + bytes_to_split],"big") for i in range(4, len(data), bytes_to_split)]
                # time_of_eve=data[i:i + bytes_to_split]
                h_excel_queue.put([plc,h,on_off_c,pq,it,time_of_eve])
                h_chat_queue.put([plc,h,pq,it,tpq,tit,time_of_eve])
            # elif data[0]==3:
            #     plc=data[1]
            #     # h=data[2]
            #     pq,it,time_of_eve = [int.from_bytes(data[i:i + bytes_to_split],"big") for i in range(2, len(data), bytes_to_split)]
            #     # time_of_eve=data[i:i + bytes_to_split]
            #     h_excel_queue.put([plc,"Total",pq,it,time_of_eve])
                # pass
            # it_excel_queue.put(plcaddr,it,)
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'DD: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'DD: Error acquired and recorded. Error: {e}')
            else: 
                print(f'DD: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)


def move_excel_it():
    write_flag=False
    while True:
        try:
            level,plc,ft,rt,it=it_excel_queue.get()
            if not os.path.isfile(filename_of_i_excel):
                wb=Workbook()
            else:
                wb=load_workbook(filename_of_i_excel)
            station=filedic[plc]
            # if station_list!=wb.sheetnames:
            #     for station in station_list:
            if not station in wb.sheetnames:
                    wb.create_sheet(station)
            ws=wb[station]
            now=datetime.datetime.now()
            date=now.strftime("%d-%m-%Y")
            time_format=now.strftime("%I.%M.%S_%p")
            # print(date,time_to_save)
            excel_dic={}
            excel_dic["DATE"]=date
            excel_dic["TIME"]=time_format
            excel_dic["SHIFT"]=get_shift(datetime.time(now.hour,now.minute,now.second))
            excel_dic["LINE"]=filedic[plc]
            # excel_dic["ALARM"]=raw_dic[err_id]
            excel_dic["FROM_TIME"]=ft
            excel_dic["IDLE_TIME_IN_MINS"]=round(it,2)
            excel_dic["LEVEL"]=level
            excel_dic["RECV_TIME"]=rt
            
            xl_headers=[]
            for i in ws[1]:
                xl_headers.append(i.value)
            mc=ws.max_column
            mr=ws.max_row
            for i in excel_dic:
                if i not in xl_headers:
                    ws.cell(1,mc+1).value=i
                    xl_headers.append(i)
                    mc=ws.max_column
            for index_i,i in enumerate(xl_headers):
                for j in excel_dic:
                    if i==j:
                        ws.cell(mr+1,index_i+1).value=excel_dic[i]
                        # ws.cell(mr+1,xl_headers.index(i)+1).value=excel_dic[i]
            while True:
                try:
                    wb.save(filename_of_i_excel)
                    print(f"EXI: Idle time Data saved successfully {time.strftime('%d-%m-%Y_%I.%M.%S_%p')} {excel_dic['LINE']}")
                except:
                    # wb.save(filename_of_excel)
                    print(f"EXI: Data not saved, Close the excel file({filename_of_i_excel}) if it is opened. Retrying to save...")
                    time.sleep(3)
                    continue
                break
            wb.close()
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'EXI: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'EXI: Error acquired and recorded. Error: {e}')
            else: 
                print(f'EXI: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)

def move_excel_h():
    # line_column,date_column,shift_column,time_h_column,time_i_column,on_off_c_column,reqiured_row=1,1,1,1,1,1,1
    reqiured_row=1
    write_flag=False
    while True:
        try:
            transfer_dic={}
            plc,h,on_off_c,pq,it,time_eve=h_excel_queue.get()
            t=datetime.datetime.fromtimestamp(time_eve)
            current_time=datetime.time(t.hour,t.minute,t.second)
            if current_time<datetime.time(A[0],A[1],A[2]):
                dp=1
            else:
                dp=0
            # plc,"Total",pq,it,time_of_eve
            date=t-datetime.timedelta(days=dp)
            date=date.strftime("%d-%m-%Y")
            # time_now=now.strftime("%I.%M.%S_%p")
            transfer_dic["LINE"]=filedic[plc]
            transfer_dic["DATE"]=date
            transfer_dic["HOUR"]=filedic[h]
            transfer_dic["SHIFT"]=get_shift(current_time)
            transfer_dic["ON/OFF_Count"]=on_off_c
            transfer_dic["PQ"]=pq
            # line_pq=filedic [plc]+"_pq"
            # line_it=filedic[plc]+"_it"
            # time_h=h+"_pq"
            # time_i=h+"_it"
            # transfer_dic[time_h]=pq
            transfer_dic["IT"]=round(it/60,2)
            
            # if not os.path.isfile(filename_of_excel):
            if not os.path.isfile(filename_of_h_excel):
                wb=Workbook()
            else:
                wb=load_workbook(filename_of_h_excel)
            station=hourly_sheet_name
            # if station_list!=wb.sheetnames:
            #     for station in station_list:
            if not station in wb.sheetnames:
                wb.create_sheet(station)
                ws=wb[station]
                for i in range(1,total_columns+1):
                    ws.cell(row=header_row_h,column=i).value=filedic[str(i)]
            ws=wb[station]        
            xl_headers=[]
            for i in ws[header_row_h]:
                xl_headers.append(i.value)
            # mc=ws.max_column
            # mr=ws.max_row
            # for i in transfer_dic:
            #     if i not in xl_headers:
            #         ws.cell(1,mc+1).value=i
            #         xl_headers.append(i)
            #         mc=ws.max_column
            date_column=int(filedic["date_column"])
            line_column=int(filedic["line_column"])
            # for i in range(1,mc+ 1):
            #     # if ws.cell(row=header_row_h,coulmn=i):
            #     v=ws.cell(row=header_row_h,column=i).value
            #     if v=="LINE":
            #         line_column=i
            #         # print(i)
            #     if v=="DATE":
            #         date_column=i
                    # print(i)
                # if v=="SHIFT":
                #     shift_column=i
                    # print(i)
                # if v==time_h:
                #     time_h_column=i
                #     # print(i)
                # if v==time_i:
                #     time_i_column=i
                # if v=="ON/OFF_Count":
                #     on_off_c_column=i
                    # print(i)
            compare_list=[transfer_dic["LINE"],transfer_dic["DATE"]]
            value_list=[]
            exist_flag=False
            h=int(h)
            for row in ws.rows:
                for cell in row:
                    value_list.append(cell.value)
                # print(type(row))
                check = all(item in value_list for item in compare_list)
                if check:
                    exist_flag=True
                    reqiured_row=cell.row
                    ws.cell(row=reqiured_row,column=h).value=transfer_dic["PQ"]
                    ws.cell(row=reqiured_row,column=h+1).value=transfer_dic["IT"]
                    if transfer_dic["SHIFT"]=="A":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countA"])).value=transfer_dic["ON/OFF_Count"]
                    elif transfer_dic["SHIFT"]=="B":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countB"])).value=transfer_dic["ON/OFF_Count"]
                    else:
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countC"])).value=transfer_dic["ON/OFF_Count"]
                    break
            mr=ws.max_row
            # print( line_column,date_column,shift_column,time_h_column,time_i_column)

            if not exist_flag:
                reqiured_row=mr+1
                ws.cell(row=reqiured_row,column=line_column).value=transfer_dic["LINE"]
                ws.cell(row=reqiured_row,column=date_column).value=transfer_dic["DATE"]
                ws.cell(row=reqiured_row,column=h).value=transfer_dic["PQ"]
                ws.cell(row=reqiured_row,column=h+1).value=transfer_dic["IT"]
                if transfer_dic["SHIFT"]=="A":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countA"])).value=transfer_dic["ON/OFF_Count"]
                elif transfer_dic["SHIFT"]=="B":
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countB"])).value=transfer_dic["ON/OFF_Count"]
                else:
                        ws.cell(row=reqiured_row,column=int(filedic["on_off_countC"])).value=transfer_dic["ON/OFF_Count"]
            raw_list=list(ws.rows)
            for t in total_dic:
                total=0
                for a in total_dic[t]:
                    v=raw_list[reqiured_row-1][a-1].value
                    if not v is None:
                        total=total+int(v)
                ws.cell(row=reqiured_row,column=t).value=total
                # if ws.cell(row=i,column=line_column).value==transfer_dic["LINE"]:
                    # if ws.cell(row=i,column=line_column).value==transfer_dic["LINE"]:
                
            # it_chat_queue.put([2,transfer_dic["LINE"],transfer_dic["DATE"],transfer_dic["SHIFT"],transfer_dic[time_h],transfer_dic[time_i]])

            # for i in xl_headers:.lo0  
            # .0   
            #     for j in transfer_dic:
            #         if i==j:
            #             ws.cell(mr+1,xl_headers.index(i)+1).value=excel_dic[i]
            while True:
                try:
                    wb.save(filename_of_h_excel)
                    print(f"EXH: Hourly and Idle time data saved successfully {time.strftime('%d-%m-%Y_%I.%M.%S_%p')} {transfer_dic['LINE']}")
                except:
                    # wb.save(filename_of_excel)
                    print(f"EXH: Data not saved, Close the excel file({filename_of_h_excel}) if it is opened. Retrying to save...")
                    time.sleep(3)
                    continue
                break
            wb.close()
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'EXH: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'EXH: Error acquired and recorded. Error: {e}')
            else: 
                print(f'EXH: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)


def	google_chat_it():
    write_flag=False
    while True:
        try:
            s1_flag,s2_flag,s3_flag=False,False,False
            level,plc,ft,rt,it=it_chat_queue.get()
            line=filedic[plc]
            line_color=filedic[line]
            if level[1] =="O":
                if level[0]==1:
                    name=level1_name
                elif level[0]==2:
                    name=level2_name
                elif level[0]==3:
                    name=level3_name
                temp_data = f'<i><b>Line Idle Alert!</b></i>\n'\
                            f'<b>Idle time crossed {round(it)} Mins &nbsp; </b> \n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line}</b>\n'\
                            f'<b>Escalation Level &nbsp; :</b> {name}\n'\
                            f'<b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}</font>\n'
                data_dir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#,
                if level[0]==2:
                        temp_data = f'<font color=\"#0000ff\"><i><b>Line Idle Alert!</b></i>\n'\
                            f'<b>Idle time crossed {round(it)} Mins &nbsp; </b> \n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line}</b>\n'\
                            f'<font color=\"#0000ff\"><b>Escalation Level &nbsp; :</b> {name}\n'\
                            f'<b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}</font>\n'
                        data_indir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#,
                        # data_indir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'<font color=\"#0000ff\">{temp_data}</font>'}}]}]}]}#, 
                elif level[0]==3:
                        temp_data = f'<font color=\"#ff0000\"><i><b>Line Idle Alert!</b></i>\n'\
                            f'<b>Idle time crossed {round(it)} Mins &nbsp; </b> \n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line}</b>\n'\
                            f'<font color=\"#ff0000\"><b>Escalation Level &nbsp; :</b> {name}\n'\
                            f'<b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}</font>\n'
                        data_indir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#,
                        # data_indir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'<font color=\"#ff0000\">{temp_data}</font>'}}]}]}]}#, 
            elif level[1] =="C":
                temp_data = f'<font color=\"#00ff00\"><i><b>Idle Time Cleared!</b></i>\n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line}</b>\n'\
                            f'<font color=\"#00ff00\"><b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}\n'\
                            f'<b>Cleared Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {rt}\n'\
                            f'<b>Idle time in Mins &nbsp; :</b> {round(it,2)}\n</font>'
                data_dir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#, 
                data_indir=data_dir
            while True:
                try:
                    if level[0]==1:
                        if not s1_flag:
                            r = requests.post(mfg_webhook_url1, data=json.dumps(data_dir))#, headers={'Content-Type': 'application/json'})
                            print("GCI: Idle time message sent to Level ",level,filedic[plc])
                            s1_flag=True
                    if level[0]==2:
                        if not s2_flag:
                            r = requests.post(mfg_webhook_url2, data=json.dumps(data_dir))#, headers={'Content-Type': 'application/json'})
                            s2_flag=True
                        if not s1_flag:
                            r = requests.post(mfg_webhook_url1, data=json.dumps(data_indir))#, headers={'Content-Type': 'application/json'})
                            s1_flag=True
                        print("GCI: Idle time message sent to Level ",level,filedic[plc])
                    if level[0]==3:
                        if not s3_flag:
                            r = requests.post(mfg_webhook_url3, data=json.dumps(data_dir))#, headers={'Content-Type': 'application/json'})
                            s3_flag=True
                        if not s2_flag:
                            r = requests.post(mfg_webhook_url2, data=json.dumps(data_indir))#, headers={'Content-Type': 'application/json'})
                            s2_flag=True
                        if not s1_flag:
                            r = requests.post(mfg_webhook_url1, data=json.dumps(data_indir))#, headers={'Content-Type': 'application/json'})
                            s1_flag=True
                        print("GCI: Idle time message sent to Level ",level,filedic[plc])
                except :
                    print("GCI: Connection Error! check internet connection. Retrying to connect...")
                    time.sleep(3)
                    continue
                break
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'GCI: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'GCI: Error acquired and recorded. Error: {e}')
            else: 
                print(f'GCI: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)

def google_chat_h():
                # h_chat_queue.put([plc,h,pq,it,time_of_eve])
    timeout_temp=None
    head_added=False
    hour_data=""
    write_flag=False
    while True:
        # bool(None)
        try:
            try:
                data_list=h_chat_queue.get(timeout=timeout_temp)
                timeout_temp=h_chat_queue_timeout
                send_msg_flag=False
            except queue.Empty:
                timeout_temp=None
                # hour_data=hour_data+f'<i><b>Development under progress, Let us know the changes need</b></i>\n'
                send_msg_flag=True
                head_added=False
                # print("empty")
            if not send_msg_flag:
                # print(a)
                plc,h,pq,it,tpq,tit,time_eve=data_list
                t=datetime.datetime.fromtimestamp(time_eve)
                current_time=datetime.time(t.hour,t.minute,t.second)
                if current_time<datetime.time(6,0,0):
                    dp=1
                else:
                    dp=0
                # plc,"Total",pq,it,time_of_eve
                date=t-datetime.timedelta(days=dp)
                date=date.strftime("%d-%m-%Y")
                # time_now=now.strftime("%I.%M.%S_%p")
                # transfer_dic["TIME"]=time_now
                shift=get_shift(current_time)
                line=filedic[plc]
                line_color=filedic[line]
                if not head_added:
                    head_data = f'<i><b><u>Hourly Intimation!</u></b></i>\n'\
                                f'<b> <u>{date} {shift} {filedic[h][:common_letters]}</u></b>\n\n'
                    hour_data=head_data
                    head_added=True
                hour_data=hour_data+f'<b><font color=\"{line_color}">Production Line &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; : {line}</b>\n'\
                                f'<b><font color=\"{line_color}">Prodution Quantity &nbsp;&nbsp; : {pq} [{tpq}]</b>\n'\
                                f'<b><font color=\"{line_color}">Machine Idle in mins : {round(it/60)} [{round(tit/60)}]</b>\n\n'
            else:
                while True:
                    try:
                        data_dir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{hour_data}'}}]}]}]}#,
                        r = requests.post(mfg_space_url, data=json.dumps(data_dir))
                        print("GCH: Google chat hourly message sent...")
                    except:
                        print("GCH: Connection Error! check internet connection. Retrying to connect...")
                        time.sleep(3)
                        continue
                    break
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'GCH: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'GCH: Error acquired and recorded. Error: {e}')
            else: 
                print(f'GCH: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)
def socket_thread():
    # soc=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
    # # print(ipaddress_of_system,port_to_listen)
    # soc.bind((ipaddress_of_system,port_to_listen))
    # soc.listen()
    
# a1=round(time.time())
# a2=round(time.time()+60)
# print(a1)
# print(a2)
# l=2

# s=3
# lb=l.to_bytes(1,'big')
# sb=s.to_bytes(1,'big')
# b1=a1.to_bytes(7,'big')
# b2=a2.to_bytes(7,'big')
# qw=True
# print(bool(0))
# qwqqq=qw.to_bytes(7,'big')
# print(qwqqq)
# print(b1)
# print(b2)
# data=lb+sb+b1+b2po    
# addr=["192.168.5.50",258]
    write_flag=False
    while True:
        try:
            soc=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
        # print(ipaddress_of_system,port_to_listen)
            soc.bind((ipaddress_of_system,port_to_listen))
            soc.listen()
            print("ST: waiting for message from client...")
            # plc_addr=input("ipaddress : ")
            # data=  
            conn,addr=soc.accept()
            pc_addr=addr[0]
            print(f"ST: connected {addr}")
        # error_data=b'\x00\x00\x70\x00'
            # print(addr[1],end=' ')
            data=conn.recv(1024)
            # print()
            data_queue.put(data)
            print(f"ST: Data received {pc_addr,data} ")
            # error_data=b'\x00\x00\x00\x00\x00\x00\x00\x01'
            # print("error_data",error_data)
            # conn.sendall(b'\x09')  
            # soc.shutdown(socket.SHUT_RDWR)   
            conn.close()
            soc.close()
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'ST: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'ST: Error acquired and recorded. Error: {e}')
            else: 
                print(f'ST: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)

    # 123nm
    # ose()
def send_mail(date,shift,table_list):
        head_list=table_list.pop(0)
        html_str=f'''<!DOCTYPE html>
        <html>
        <table style="font-family:  Arial, Helvetica,sans-serif;
        border: 1px solid #ddd;
        border-collapse:collapse;
        width: 100%;
        margin: 25px 0;
        font-size: 1.0em;
        min-width: 400px;
        border-radius: 10px 10px 0 0;
        overflow: hidden;
        box-shadow: 0 0 20px rgba(0, 0, 0, 0.8);">
        <tr>
            <th style="padding:12px;text-align: center;background-color: #04AA6D;color: white;border: 1px solid #ddd;"colspan="{len(head_list)}">DATE : {date} SHIFT: {shift}</th>
        </tr>'''
        # print(len(head_list))
        html_str=html_str+'<tr>'
        # print(head_list)
        # -ms-writing-mode: tb-rl; -webkit-writing-mode: vertical-rl; writing-mode: vertical-rl; transform: rotate(180deg); white-space: nowrap;
        for i in head_list:
            # print(i)
            html_str=html_str+f'<th style="padding:12px;text-align: center;background-color: #04AA6D;color: white;border: 1px solid #ddd;-ms-writing-mode: tb-rl; -webkit-writing-mode: vertical-rl; writing-mode: vertical-rl; transform: rotate(180deg); white-space: nowrap;">{i}</th>'
        html_str=html_str+'</tr>'
        for sub_list in table_list:
            html_str=html_str+'<tr>'
            for index_i,i in enumerate(sub_list):
                if index_i%2:
                    html_str=html_str+f'<td style="border: 1px solid #ddd;padding: 11px;text-align: center;">{i}</td>'
                else:
                    html_str=html_str+f'<td style="border: 1px solid #ddd;padding: 11px;text-align: center;background-color: #e6e6e6;">{i}</td>'
            html_str=html_str+'</tr>'
        # transform: rotate(180deg);
        html_str=html_str+'''</table></html>'''
        # print(html_str)
#         return html_str

# def send_mail(dat):
        body=html_str
        message = MIMEMultipart()
        message['Subject'] = 'Production report'
        message['From'] = 'osdple@gmail.com'
        # emails = ['p.thirunavukkarasu@ranegroup.com','osdple@gmail.com']
        message['To'] = ', '.join( list_of_mails ) 
        # message['To'] = 'p.thirunavukkarasu@ranegroup.com,osdple@gmail.com'

        body_content = body
        message.attach(MIMEText(body_content, "html"))
        msg_body = message.as_string()
        # msg_body = body
        # time.sleep()
        while True:
            try:
                server = SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(message['From'], 'osdple@123')
                server.sendmail(message['From'], list_of_mails, msg_body)
                server.quit()
                print("SM: Summary mail sent...")
            except Exception as e:
                print("SM: Unable to sent mail, ",e)
                time.sleep(3)
                continue
            break


def trig_mail_condent(shift):
    time.sleep(mail_trig_delay)
    table_list=[]
    required_row_list=[]
    required_column_list=[]
    no_data=True
    time_now=datetime.datetime.now()
    # date=time_now.strftime("%d/%m/%Y")
    if shift=="C":
            dp=1
    else:
            dp=0
        # plc,"Total",pq,it,time_of_eve
    time_now=time_now-datetime.timedelta(days=dp)
    date_today=time_now.strftime("%d-%m-%Y")
    summary_list=summary_dic[shift]
    table_list.append(summary_list)
    if os.path.isfile(filename_of_h_excel):
        wb=load_workbook(filename_of_h_excel)
        ws=wb[hourly_sheet_name]
        xl_headers=[]
        for i in ws[header_row_h]:
                xl_headers.append(i.value)
        date_today
        mr=ws.max_row
        for i in range(1,mr+1):
            if ws.cell(row=i,column=int(filedic["date_column"])).value==date_today:
                required_row_list.append(i)
        for i in summary_list:
            required_column_list.append(xl_headers.index(i)+1)

        raw_list=list(ws.rows)
        for r in required_row_list:
            no_data=False
            row_list=[]
            for c in required_column_list:
                row_list.append(raw_list[r-1][c-1].value)
            table_list.append(row_list)
        # shift=shift+" (Development under progress, Let us know the changes need)"
        if no_data:
            print("TMC: As no data for the shift, mail not sent")
        else:
            send_mail(date_today,shift,table_list)
        wb.close()
    else:
        print(f"TMC: As no {filename_of_h_excel} found, mail not sent")


def mail_trigger():
    try:
        schedule.every().day.at(shiftA_start).do(trig_mail_condent,"C")
        schedule.every().day.at(shiftB_start).do(trig_mail_condent,"A")
        schedule.every().day.at(shiftC_start).do(trig_mail_condent,"B")
        while True:
            schedule.run_pending()
            time.sleep(scheduler_delay)
    except Exception as e:
        error_register_queue.put(f'MAT: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
        print(f'MAT: Error acquired and recorded. Error: {e}')
        time.sleep(error_wait)

data_dec = threading.Thread(target=data_decode,daemon=True)
data_dec.start()
print("MT: Data_decode thread started")
move_ex_i = threading.Thread(target=move_excel_it,daemon=True)
move_ex_i.start()
print("MT: Move_excel_i thread started")
move_ex_h = threading.Thread(target=move_excel_h,daemon=True)
move_ex_h.start()
print("MT: Move_excel_h thread started")
google_ct_i = threading.Thread(target=google_chat_it,daemon=True)
google_ct_i.start()
print("MT: Google_chat_it thread started")
google_ct_h = threading.Thread(target=google_chat_h,daemon=True)
google_ct_h.start()
print("MT: Google_chat_h thread started")
mail_trig = threading.Thread(target=mail_trigger,daemon=True)
mail_trig.start()
print("MT: Mail_trigger thread started")
socket_th = threading.Thread(target=socket_thread,daemon=True)
socket_th.start()
print("MT: Socket_thread thread started")
error_reg = threading.Thread(target=error_register,daemon=True)
error_reg.start()
print(f"MT: Error_register thread started")

while True:
    time.sleep(60)