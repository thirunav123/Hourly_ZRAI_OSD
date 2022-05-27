import datetime,os
from openpyxl import Workbook,load_workbook
total_file= open("total.txt", "r")
filedic={}
# list(map(int,shiftA_start.strip().split(":"))
for line in total_file:
    file_data=line.strip().split('===')
    a=int(file_data[0])
    # b=.strip().split(",")
    b=map(int,file_data[1].strip().split(","))
    filedic[a]=b
total_file.close()
print(filedic)
station="st"
wb=Workbook()
wb.create_sheet(station)
ws=wb[station]
for i in range(68):
    # if i%2==0:
        ws.cell(row=1,column=i+1).value=1
raw_list=list(ws.rows)
r=1
for t in filedic:
    total=0
    for a in filedic[t]:
        v=raw_list[r-1][a-1].value
        if not v is None:
            total=total+v
    ws.cell(row=r,column=t).value=total
wb.save("total.xlsx")
# wb.close()
# wob=load_workbook("a.xlsx",data_only=True)
# wos=wob[station]
# print(type(int(wos["A1"].value)))

