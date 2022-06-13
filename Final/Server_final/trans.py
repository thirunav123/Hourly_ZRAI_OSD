import os,socket,queue,datetime,time,threading,requests,json,schedule,copy
from openpyxl import Workbook,load_workbook
from smtplib import SMTP
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

setting_file=open('server_settings.txt','r')
print("MT: I am a Server")
filedic={}
for line in setting_file:
    file_data=line.strip().split('===')
    a=file_data[0]
    b=file_data[1]
    filedic[a]=b
setting_file.close()
ipaddress_of_system=filedic.pop('ipaddress_of_system')
port_to_listen=int(filedic.pop('port_to_listen'))
shiftA_start=filedic.pop('shiftA_start_time')
shiftB_start=filedic.pop('shiftB_start_time')
shiftC_start=filedic.pop('shiftC_start_time')
filename_of_i_excel=filedic.pop('filename_of_idle_time_excel_sheet')
filename_of_h_excel=filedic.pop('filename_of_hourly_excel_sheet')
hourly_list=filedic.pop('hourly_intimate').split(",")
columns_before_pq_and_it_in_excel=int(filedic.pop('columns_before_pq_and_it_in_excel'))
mfg_webhook_url1=filedic.pop('mfg_webhook_url_level1')
mfg_webhook_url2=filedic.pop('mfg_webhook_url_level2')
mfg_webhook_url3=filedic.pop('mfg_webhook_url_level3')
level1_name=filedic.pop('level1_name')
level2_name=filedic.pop('level2_name')
level3_name=filedic.pop('level3_name')
line_no_list=filedic.pop('LINE_NUMBERS').split(",")
hourly_sheet_name=filedic.pop('hourly_sheet_name')
header_row_h=int(filedic.pop('header_row_hourly'))
total_columns=int(filedic.pop('total_columns'))
mfg_space_url=filedic.pop('mfg_space_url')
scheduler_delay=int(filedic.pop("scheduler_delay_in_milliseconds"))/1000
error_wait=int(filedic.pop("error_wait_in_milliseconds"))/1000
mail_trig_delay=int(filedic.pop("mail_trigger_delay_after_next_shift_start_in_seconds"))
common_letters=int(filedic.pop("common_letters_count_from_begin_in_pq_and_it_columns"))
json_file_name=filedic.pop('json_file_name_for_store_values')
dict_backup_time=int(filedic.pop('dict_backup_time_in_seconds'))
idle_time_calc_delay=int(filedic.pop("idle_time_calc_delay_in_seconds"))
prior_time=int(filedic.pop('prior_time_for_hourly_trigger_in_seconds'))
end_timing_pos=filedic.pop('end_timing_index_positins')
mail_user_name=filedic.pop("mail_user_name")
app_password=filedic.pop('app_password')
print("MT: Settings file read")

total_file= open("total.txt", "r")
mail_file = open("mail.txt", "r")
summary_a_file= open("summary_A.txt", "r")
summary_b_file= open("summary_B.txt", "r")
summary_c_file= open("summary_C.txt", "r")

total_dic={}
list_of_mails= []
summary_a_list= []
summary_b_list= []
summary_c_list= []
for line in total_file:
    file_data=line.strip().split('===')
    a=int(file_data[0])
    b=list(map(int,file_data[1].strip().split(",")))
    total_dic[a]=b
for line in mail_file:
  list_of_mails.append(line.strip())
for line in summary_a_file:
  summary_a_list.append(line.strip())
for line in summary_b_file:
  summary_b_list.append(line.strip())
for line in summary_c_file:
  summary_c_list.append(line.strip())

total_file.close()
mail_file.close()
summary_a_file.close()
summary_b_file.close()
summary_c_file.close()
summary_dic={"A":summary_a_list,"B":summary_b_list,"C":summary_c_list}
print("MT: Total, Mail, Summary_A, Summary_B, Summary_C file read")
A=list(map(int,shiftA_start.strip().split(":")))
B=list(map(int,shiftB_start.strip().split(":")))
C=list(map(int,shiftC_start.strip().split(":")))
end_timing_pos_list=list(map(int,end_timing_pos.strip().split(",")))
data_queue=queue.Queue()
it_excel_queue=queue.Queue()
it_chat_queue=queue.Queue()
h_chat_queue=queue.Queue()
h_excel_queue=queue.Queue()
error_register_queue=queue.Queue()
sub_dict_default={"opq":0,"oit":0,"npq":0,"nit":0,"ogit":0,"ini_t":round(time.time()),"ooc":0,"hrf":False,"srf":False,"mac_s":False,"ef":False,"l1f":False,"l2f":False,"l3f":False}
if not os.path.isfile(json_file_name):
    json_dict={}
    for line_no in line_no_list:
        json_dict[line_no]=sub_dict_default.copy()
    json_file=open(json_file_name,"w")
    json.dump(json_dict, json_file)
    json_file.close()
    print("MT: Json file created")
json_read_file=open(json_file_name,"r")
loaded_str=json_read_file.read()
json_read_file.close()
loaded_dict=json.loads(loaded_str)
# print(loaded_dict)
new_line_added,line_removed=False,False
for line_no in line_no_list:
    if line_no not in loaded_dict.keys():
        # opq: old production quantity,npq: new production quantity,oit: old idle time
        # nit: new idle time,ogit: on going idletime,ini_t:initial time,ooc: on off count, hrf: hour reset flag, srf: shift reset flag
        # ef: entry falg, l1f: Level 1 flag,l2f: Level 2 flag,l3f: Level 3 flag
        loaded_dict[line_no]=sub_dict_default
        new_line_added=True
line_remove_list=[]
for line_no in loaded_dict:
    if line_no not in line_no_list:
        line_remove_list.append(line_no)
        line_removed=True
if line_removed:
    for line_no in line_remove_list:
        loaded_dict.pop(line_no)
if new_line_added or line_removed:
    json_file=open(json_file_name,"w")
    json.dump(loaded_dict, json_file)
    json_file.close()
    if new_line_added:
        print("MT: New line added in Json file")
    if line_removed:
        print("MT: Existing line removed in Json file")
line_data_dict={}
for line_no in line_no_list:
    details_list=filedic.pop(line_no).split(",")
    line_name=details_list.pop(0)
    line_color=details_list.pop()
    for i in range(len(details_list)):
        details_list[i]=int(details_list[i])
    sub_dict={"l_n":line_name,"mt":details_list[0],"l1_t":details_list[1],"l2_t":details_list[2],"l3_t":details_list[3],"l_c":line_color}
    line_data_dict[line_no]=sub_dict
unique_no_in_bytes=None

def get_shift(ct):
    startA=datetime.time(A[0],A[1],A[2])
    startB=datetime.time(B[0],B[1],B[2])
    startC=datetime.time(C[0],C[1],C[2])
    if startA<ct<startB:
        return 'A'
    if startB<ct<startC:
        return 'B'
    else:
        return 'C'

def error_register():
    while True:
        err=error_register_queue.get()
        with open('readme.txt', 'a') as f:
            f.write(err)
        f.close()

def dict_backup():
    while True:
        json_file=open(json_file_name,"w")
        json.dump(loaded_dict, json_file)
        json_file.close()
        time.sleep(dict_backup_time)
    
def hourly(hour):
    try:
        temp_dict=copy.deepcopy(loaded_dict)
        h_excel_queue.put([hour,temp_dict])
        h_chat_queue.put([hour,temp_dict])
        if hour in end_timing_pos_list:
            for line_no in loaded_dict:
                loaded_dict[line_no]["srf"]=True
        else:
            for line_no in loaded_dict:
                loaded_dict[line_no]["hrf"]=True
        # del temp_dict
        print("H: Hourly message triggered   ",filedic[hour][:common_letters])

    except Exception as e:
        error_register_queue.put(f'H: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
        print(f'H: Error acquired and recorded. Error: {e}')
        time.sleep(error_wait)
def socket_thread():
    global unique_no_in_bytes
    write_flag=False
    while True:
        if not unique_no_in_bytes is None:
            break
        time.sleep(1)
    while True:
        try:
            soc=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
            soc.bind((ipaddress_of_system,port_to_listen))
            soc.listen()
            conn,addr=soc.accept()
            pc_addr=addr[0]
            data=conn.recv(1024)
            conn.sendall(unique_no_in_bytes)
            if data[0]==1:
                pass
            else:
                data_queue.put(data)
            print(f"ST: Data received {pc_addr, data} ")   
            conn.close()
            soc.close()
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'ST: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'ST: Error acquired and recorded. Error: {e}')
            else: 
                print(f'ST: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)

def data_decode():
    write_flag=False
    while True:
        try:
            data=data_queue.get()
            data=bytearray(data)
            line_no=str(data.pop(0))
            count=int.from_bytes(data,'big')
            if count==65535:
                loaded_dict[line_no]["ooc"]=loaded_dict[line_no]["ooc"]+1
                loaded_dict[line_no]["mac_s"]=False
            else:
                loaded_dict[line_no]["npq"]=int.from_bytes(data,'big')
                loaded_dict[line_no]["mac_s"]=True
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'DD: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'DD: Error acquired and recorded. Error: {e}')
            else: 
                print(f'DD: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)

def idle_time_calc(line_no):
    # entry_flag,level1_flag,level2_flag,level3_flag=False,False,False,False
    initial_time=loaded_dict[line_no]["ini_t"]
    og_int_time=initial_time
    write_flag=False
    pq_count=loaded_dict[line_no]["npq"]
    while True:
        try:
                current_time=round(time.time())
                if loaded_dict[line_no]["hrf"]:
                    loaded_dict[line_no]["opq"]=loaded_dict[line_no]["npq"]
                    loaded_dict[line_no]["oit"]=loaded_dict[line_no]["nit"]+loaded_dict[line_no]["ogit"]
                    loaded_dict[line_no]["hrf"]=False
                if loaded_dict[line_no]["srf"]:
                    loaded_dict[line_no]["opq"]=0
                    loaded_dict[line_no]["npq"]=0
                    loaded_dict[line_no]["oit"]=0
                    loaded_dict[line_no]["nit"]=0
                    loaded_dict[line_no]["ogit"]=0
                    loaded_dict[line_no]["srf"]=False
                    og_int_time=current_time
                if pq_count==loaded_dict[line_no]["npq"]:
                        if initial_time+line_data_dict[line_no]["mt"]<current_time:
                            if not loaded_dict[line_no]["ef"]:
                                from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                idle_time=round((current_time-initial_time)/60,2)
                                it_excel_queue.put(["Instance",line_data_dict[line_no]["l_n"],from_time,trig_time,idle_time])
                                loaded_dict[line_no]["ef"]=True
                                print(f'ITC: Idle time started, Entry registered {line_data_dict[line_no]["l_n"]}')
                            if initial_time+line_data_dict[line_no]["l1_t"]<current_time:
                                if not loaded_dict[line_no]["l1f"]:
                                    if loaded_dict[line_no]["mac_s"]:
                                        from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                        trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                        idle_time=round((current_time-initial_time)/60,2)
                                        it_excel_queue.put([1,line_data_dict[line_no]["l_n"],from_time,trig_time,idle_time])
                                        it_chat_queue.put([[1,"O"],line_no,from_time,trig_time,idle_time])
                                        loaded_dict[line_no]["l1f"]=True
                                        print(f'ITC: Level1 {line_data_dict[line_no]["l_n"]}')
                                if initial_time+line_data_dict[line_no]["l2_t"]<current_time:
                                    if not loaded_dict[line_no]["l2f"]:
                                        if loaded_dict[line_no]["mac_s"]:
                                            from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                            trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                            idle_time=round((current_time-initial_time)/60,2)
                                            it_excel_queue.put([2,line_data_dict[line_no]["l_n"],from_time,trig_time,idle_time])
                                            it_chat_queue.put([[2,"O"],line_no,from_time,trig_time,idle_time])
                                            loaded_dict[line_no]["l2f"]=True
                                            print(f'ITC: Level2 {line_data_dict[line_no]["l_n"]}')
                                    if initial_time+line_data_dict[line_no]["l3_t"]<current_time:
                                        if not loaded_dict[line_no]["l3f"]:
                                            if loaded_dict[line_no]["mac_s"]:
                                                from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                                trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                                idle_time=round((current_time-initial_time)/60,2)
                                                it_excel_queue.put([3,line_data_dict[line_no]["l_n"],from_time,trig_time,idle_time])
                                                it_chat_queue.put([[3,"O"],line_no,from_time,trig_time,idle_time])
                                                loaded_dict[line_no]["l3f"]=True
                                                print(f'ITC: Level3 {line_data_dict[line_no]["l_n"]}')
                        if loaded_dict[line_no]["ef"] or loaded_dict[line_no]["l1f"] or loaded_dict[line_no]["l2f"] or loaded_dict[line_no]["l3f"]:
                            loaded_dict[line_no]["ogit"]=current_time-og_int_time
                else:
                            if loaded_dict[line_no]["ef"] or loaded_dict[line_no]["l1f"] or loaded_dict[line_no]["l2f"] or loaded_dict[line_no]["l3f"]:
                                loaded_dict[line_no]["nit"]=loaded_dict[line_no]["nit"]+loaded_dict[line_no]["ogit"]
                                from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                idle_time=round((current_time-initial_time)/60,2)
                                it_excel_queue.put(["Cleared",line_data_dict[line_no]["l_n"],from_time,trig_time,idle_time])
                                loaded_dict[line_no]["ogit"]=0
                            if loaded_dict[line_no]["l3f"]:
                                from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                idle_time=round((current_time-initial_time)/60,2)
                                it_chat_queue.put([[3,"C"],line_no,from_time,trig_time,idle_time])
                                print(f'ITC: Level 3 Idle time cleared {line_data_dict[line_no]["l_n"]}')
                            elif loaded_dict[line_no]["l2f"]:
                                from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                idle_time=round((current_time-initial_time)/60,2)
                                it_chat_queue.put([[2,"C"],line_no,from_time,trig_time,idle_time])
                                print(f'ITC: Level 2 Idle time cleared {line_data_dict[line_no]["l_n"]}')
                            elif loaded_dict[line_no]["l1f"]:
                                from_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p",time.localtime(initial_time))
                                trig_time=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
                                idle_time=round((current_time-initial_time)/60,2)
                                it_chat_queue.put([[1,"C"],line_no,from_time,trig_time,idle_time])
                                print(f'ITC: Level 1 Idle time cleared {line_data_dict[line_no]["l_n"]}')
                            elif loaded_dict[line_no]["ef"]:
                                print(f'ITC: Instance Idle time cleared {line_data_dict[line_no]["l_n"]}')
                            loaded_dict[line_no]["ef"],loaded_dict[line_no]["l1f"],loaded_dict[line_no]["l2f"],loaded_dict[line_no]["l3f"]=False,False,False,False
                            pq_count=loaded_dict[line_no]["npq"]
                            initial_time=round(time.time())
                            og_int_time=initial_time
                            loaded_dict[line_no]["ini_t"]=initial_time
                time.sleep(idle_time_calc_delay)
                write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'ITC: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'ITC: Error acquired and recorded. Error: {e}')
            else: 
                print(f'ITC: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)        

def move_excel_it():
    write_flag=False
    while True:
        try:
            level,line,ft,tt,it=it_excel_queue.get()
            if not os.path.isfile(filename_of_i_excel):
                wb=Workbook()
            else:
                wb=load_workbook(filename_of_i_excel)
            if not line in wb.sheetnames:
                    wb.create_sheet(line)
            ws=wb[line]
            now=datetime.datetime.now()
            date=now.strftime("%d-%m-%Y")
            excel_dic={}
            excel_dic["DATE"]=date
            excel_dic["SHIFT"]=get_shift(datetime.time(now.hour,now.minute,now.second))
            excel_dic["FROM_TIME"]=ft
            excel_dic["IDLE_TIME_IN_MINS"]=it
            excel_dic["LEVEL"]=level
            excel_dic["TRIGGER_TIME"]=tt
            
            xl_headers=[]
            for i in ws[1]:
                xl_headers.append(i.value)
            mc=ws.max_column
            mr=ws.max_row
            for i in excel_dic:
                if i not in xl_headers:
                    ws.cell(1,mc+1).value=i
                    xl_headers.append(i)
                    mc=ws.max_column
            for index_i,i in enumerate(xl_headers):
                for j in excel_dic:
                    if i==j:
                        ws.cell(mr+1,index_i+1).value=excel_dic[i]
            while True:
                try:
                    wb.save(filename_of_i_excel)
                    print(f"EXI: Idle time Data saved successfully {time.strftime('%d-%m-%Y_%I.%M.%S_%p')} {line}")
                except Exception as e:
                    print(f"EXI: Data not saved, Close the excel file({filename_of_i_excel}) if it is opened. Retrying to save...\n Error: {e}")
                    time.sleep(3)
                    continue
                break
            wb.close()
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'EXI: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'EXI: Error acquired and recorded. Error: {e}')
            else: 
                print(f'EXI: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)


def move_excel_h():
    reqiured_row=1
    write_flag=False
    while True:
        try:
            hour,dict_ex=h_excel_queue.get()
            station=hourly_sheet_name
            if not os.path.isfile(filename_of_h_excel):
                wb=Workbook()
            else:
                wb=load_workbook(filename_of_h_excel)
            if not station in wb.sheetnames:
                wb.create_sheet(station)
                ws=wb[station]
                for i in range(1,total_columns+1):
                    ws.cell(row=header_row_h,column=i).value=filedic[str(i)]
            ws=wb[station]        
            xl_headers=[]
            for i in ws[header_row_h]:
                xl_headers.append(i.value)
            date_column=int(filedic["date_column"])
            line_column=int(filedic["line_column"])
            time_eve=round(time.time())-prior_time
            t=datetime.datetime.fromtimestamp(time_eve)
            current_time=datetime.time(t.hour,t.minute,t.second)
            if current_time<datetime.time(A[0],A[1],A[2]):
                dp=1
            else:
                dp=0
            date=t-datetime.timedelta(days=dp)
            date=date.strftime("%d-%m-%Y")
            transfer_dic={}
            transfer_dic["DATE"]=date
            transfer_dic["SHIFT"]=get_shift(current_time)
            transfer_dic["HOUR"]=filedic[hour]
            for line_no in dict_ex:
                transfer_dic["LINE"]=line_data_dict[line_no]["l_n"]
                transfer_dic["ON/OFF_Count"]=dict_ex[line_no]["ooc"]
                transfer_dic["PQ"]=dict_ex[line_no]["npq"]-dict_ex[line_no]["opq"]
                transfer_dic["IT"]=round((dict_ex[line_no]["nit"]+dict_ex[line_no]["ogit"]-dict_ex[line_no]["oit"])/60)
                compare_list=[transfer_dic["LINE"],transfer_dic["DATE"]]
                exist_flag=False
                h=int(hour)
                for row in ws.rows:
                    value_list=[]
                    for cell in row:
                        value_list.append(cell.value)
                    check = all(item in value_list for item in compare_list)
                    if check:
                        exist_flag=True
                        reqiured_row=cell.row
                        ws.cell(row=reqiured_row,column=h).value=transfer_dic["PQ"]
                        ws.cell(row=reqiured_row,column=h+1).value=transfer_dic["IT"]
                        if transfer_dic["SHIFT"]=="A":
                            ws.cell(row=reqiured_row,column=int(filedic["on_off_countA"])).value=transfer_dic["ON/OFF_Count"]
                        elif transfer_dic["SHIFT"]=="B":
                            ws.cell(row=reqiured_row,column=int(filedic["on_off_countB"])).value=transfer_dic["ON/OFF_Count"]
                        else:
                            ws.cell(row=reqiured_row,column=int(filedic["on_off_countC"])).value=transfer_dic["ON/OFF_Count"]
                        break
                mr=ws.max_row

                if not exist_flag:
                    reqiured_row=mr+1
                    ws.cell(row=reqiured_row,column=line_column).value=transfer_dic["LINE"]
                    ws.cell(row=reqiured_row,column=date_column).value=transfer_dic["DATE"]
                    ws.cell(row=reqiured_row,column=h).value=transfer_dic["PQ"]
                    ws.cell(row=reqiured_row,column=h+1).value=transfer_dic["IT"]
                    if transfer_dic["SHIFT"]=="A":
                            ws.cell(row=reqiured_row,column=int(filedic["on_off_countA"])).value=transfer_dic["ON/OFF_Count"]
                    elif transfer_dic["SHIFT"]=="B":
                            ws.cell(row=reqiured_row,column=int(filedic["on_off_countB"])).value=transfer_dic["ON/OFF_Count"]
                    else:
                            ws.cell(row=reqiured_row,column=int(filedic["on_off_countC"])).value=transfer_dic["ON/OFF_Count"]
                raw_list=list(ws.rows)
                for t in total_dic:
                    total=0
                    for a in total_dic[t]:
                        v=raw_list[reqiured_row-1][a-1].value
                        if not v is None:
                            total=total+float(v)
                    ws.cell(row=reqiured_row,column=t).value=total
            while True:
                try:
                    wb.save(filename_of_h_excel)
                    print(f"EXH: Hourly and Idle time data saved successfully {time.strftime('%d-%m-%Y_%I.%M.%S_%p')} {filedic[hour][:common_letters]}")
                except Exception as e:
                    print(f"EXH: Data not saved, Close the excel file({filename_of_h_excel}) if it is opened. Retrying to save...\n Error: {e}")
                    time.sleep(3)
                    continue
                break
            wb.close()
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'EXH: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'EXH: Error acquired and recorded. Error: {e}')
            else: 
                print(f'EXH: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)


def	google_chat_it():
    write_flag=False
    while True:
        try:
            s1_flag,s2_flag,s3_flag=False,False,False
            level,line_no,ft,tt,it=it_chat_queue.get()
            line_name=line_data_dict[line_no]["l_n"]
            line_color=line_data_dict[line_no]["l_c"]
            if level[1] =="O":
                name=level1_name
                temp_data = f'<i><b>Line Idle Alert!</b></i>\n'\
                            f'<b>Idle time crossed {round(it)} Mins &nbsp; </b> \n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line_name}</b>\n'\
                            f'<b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}</font>\n'
                data_dir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#,
                if level[0]==2:
                        name=level2_name
                        temp_data = f'<font color=\"#0000ff\"><i><b>Line Idle Alert!</b></i>\n'\
                            f'<b>Idle time crossed {round(it)} Mins &nbsp; </b> \n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line_name}</b>\n'\
                            f'<font color=\"#0000ff\"><b>Escalation Level &nbsp; :</b> {name}\n'\
                            f'<b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}</font>\n'
                        data_indir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#,
                elif level[0]==3:
                        name=level3_name
                        temp_data = f'<font color=\"#ff0000\"><i><b>Line Idle Alert!</b></i>\n'\
                            f'<b>Idle time crossed {round(it)} Mins &nbsp; </b> \n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line_name}</b>\n'\
                            f'<font color=\"#ff0000\"><b>Escalation Level &nbsp; :</b> {name}\n'\
                            f'<b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}</font>\n'
                        data_indir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#,
            elif level[1] =="C":
                temp_data = f'<font color=\"#00ff00\"><i><b>Idle Time Cleared!</b></i>\n'\
                            f'<b>Production Line &nbsp;&nbsp; : <font color=\"{line_color}">{line_name}</b>\n'\
                            f'<font color=\"#00ff00\"><b>From Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {ft}\n'\
                            f'<b>Cleared Time &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</b> {tt}\n'\
                            f'<b>Idle time in Mins &nbsp; :</b> {it}\n</font>'
                data_dir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{temp_data}'}}]}]}]}#, 
                data_indir=data_dir
            while True:
                try:
                    if level[0]==1:
                        if not s1_flag:
                            r = requests.post(mfg_webhook_url1, data=json.dumps(data_dir))
                            print("GCI: Idle time message sent to Level ",level,line_name)
                            s1_flag=True
                    if level[0]==2:
                        if not s2_flag:
                            r = requests.post(mfg_webhook_url2, data=json.dumps(data_dir))
                            s2_flag=True
                        if not s1_flag:
                            r = requests.post(mfg_webhook_url1, data=json.dumps(data_indir))
                            s1_flag=True
                        print("GCI: Idle time message sent to Level ",level,line_name)
                    if level[0]==3:
                        if not s3_flag:
                            r = requests.post(mfg_webhook_url3, data=json.dumps(data_dir))
                            s3_flag=True
                        if not s2_flag:
                            r = requests.post(mfg_webhook_url2, data=json.dumps(data_indir))
                            s2_flag=True
                        if not s1_flag:
                            r = requests.post(mfg_webhook_url1, data=json.dumps(data_indir))
                            s1_flag=True
                        print("GCI: Idle time message sent to Level ",level,line_name)
                except Exception as e:
                    print(f"GCI: Connection Error! check internet connection. Retrying to connect... \n Error: {e}")
                    time.sleep(3)
                    continue
                break
            write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'GCI: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'GCI: Error acquired and recorded. Error: {e}')
            else: 
                print(f'GCI: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)

def google_chat_h():
    hour_data=""
    write_flag=False
    while True:
        try:
                hour,dict_ct=h_chat_queue.get()
                time_eve=round(time.time())-prior_time
                t=datetime.datetime.fromtimestamp(time_eve)
                current_time=datetime.time(t.hour,t.minute,t.second)
                if current_time<datetime.time(A[0],A[1],A[2]):
                    dp=1
                else:
                    dp=0
                date=t-datetime.timedelta(days=dp)
                date=date.strftime("%d-%m-%Y")
                shift=get_shift(current_time)
                hour_data = f'<i><b><u>Hourly Intimation!</u></b></i>\n'\
                            f'<b> <u>{date} {shift} {filedic[hour][:common_letters]}</u></b>\n\n'
                for line_no in dict_ct:
                    line_name=line_data_dict[line_no]["l_n"]
                    line_color=line_data_dict[line_no]["l_c"]
                    pq=dict_ct[line_no]["npq"]-dict_ct[line_no]["opq"]
                    tpq=dict_ct[line_no]["npq"]
                    it=round((dict_ct[line_no]["nit"]+dict_ct[line_no]["ogit"]-dict_ct[line_no]["oit"])/60)
                    tit=round((dict_ct[line_no]["nit"]+dict_ct[line_no]["ogit"])/60)
                    hour_data=hour_data+f'<b><font color=\"{line_color}">Production Line &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; : {line_name}</b>\n'\
                                f'<b><font color=\"{line_color}">Production Quantity &nbsp; : {pq} [{tpq}]</b>\n'\
                                f'<b><font color=\"{line_color}">Machine Idle in mins : {it} [{tit}]</b>\n\n'
                while True:
                    try:
                        data_dir = {"cards": [{"sections":[{"widgets":[{"textParagraph":{ 'text':f'{hour_data}'}}]}]}]}#,
                        r = requests.post(mfg_space_url, data=json.dumps(data_dir))
                        print(f"GCH: Google chat hourly message sent... {date} {shift} {filedic[hour][:common_letters]}")
                    except Exception as e:
                        print(f"GCH: Connection Error! check internet connection. Retrying to connect... \n Error: {e}")
                        time.sleep(3)
                        continue
                    break
                write_flag=False
        except Exception as e:
            if not write_flag:
                error_register_queue.put(f'GCH: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
                write_flag=True
                print(f'GCH: Error acquired and recorded. Error: {e}')
            else: 
                print(f'GCH: Error acquired and already recorded. Error: {e}')
            time.sleep(error_wait)

def send_mail(date,shift,table_list):
        head_list=table_list.pop(0)
        list_len=len(head_list)
        if shift=="A":
            table_head_color="#ff5500"
        if shift=="B":
            table_head_color="#04AA6D"
        else:
            table_head_color="#5500ff"
        html_str=f'''<!DOCTYPE html>
        <html><body>
        <table style="font-family:  Arial, Helvetica,sans-serif;
        border: 1px solid #ddd;
        border-collapse:collapse;
        width: 100%;
        margin: 25px 0;
        font-size: 1.0em;
        min-width: 400px;
        border-radius: 10px 10px 0 0;
        overflow: hidden;
        box-shadow: 0 0 20px rgba(0, 0, 0, 0.8);">
        <tr>
            <th style="padding:12px;text-align: center;background-color: {table_head_color};color: white;border: 1px solid #ddd;"colspan="{list_len}">DATE : {date} SHIFT: {shift}</th>
        </tr>'''
        html_str=html_str+'<tr>'
        for i in range(list_len):
            if 0<i<=(list_len-3):
                if i%2: 
                    html_str=html_str+f'<th style="padding:12px;text-align: center;background-color: {table_head_color};color: white;border: 1px solid #ddd;white-space: nowrap;"colspan="2">{head_list[i][:common_letters]}</th>'
                else:
                    pass
            elif i>=(list_len-3):
                html_str=html_str+f'<th style="padding:12px;text-align: center;background-color: {table_head_color};color: white;border: 1px solid #ddd;white-space: nowrap;"colspan="2">{head_list[i][:5]}</th>'
                break
            elif i==0:
                html_str=html_str+f'<th style="padding:12px;text-align: center;background-color: {table_head_color};color: white;border: 1px solid #ddd;white-space: nowrap;"rowspan="2">{head_list[i]}</th>'
        html_str=html_str+'</tr>'
        html_str=html_str+'<tr>'
        for i in range(list_len):
            if i>0:
                if i%2:
                    html_str=html_str+f'<th style="padding:12px;text-align: center;background-color: {table_head_color};color: white;border: 1px solid #ddd;white-space: nowrap;">{head_list[i][-2:]}</th>'
                else:
                    html_str=html_str+f'<th style="padding:12px;text-align: center;background-color: {table_head_color};color: white;border: 1px solid #ddd;white-space: nowrap;">{head_list[i][-2:]}</th>'

        html_str=html_str+'</tr>'
        for sub_list in table_list:
            html_str=html_str+'<tr>'
            for index_i,i in enumerate(sub_list):
                if index_i==(list_len-1):
                    html_str=html_str+f'<td style="border: 1px solid #ddd;padding: 11px;text-align: center;background-color: #e6e6e6;"><b>{i}</b></td>'
                elif index_i==(list_len-2):
                    html_str=html_str+f'<td style="border: 1px solid #ddd;padding: 11px;text-align: center;"><b>{i}</b></td>'
                elif index_i%2:
                    html_str=html_str+f'<td style="border: 1px solid #ddd;padding: 11px;text-align: center;">{i}</td>'
                else:
                    html_str=html_str+f'<td style="border: 1px solid #ddd;padding: 11px;text-align: center;background-color: #e6e6e6;">{i}</td>'
            html_str=html_str+'</tr>'
        html_str=html_str+'</table>'
        html_str=html_str+'<p>* PQ : Production Quantity in Numbers</p><p>* IT: Idle time in Minutes</p>'
        html_str=html_str+'</body></html>'
        body=html_str
        message = MIMEMultipart()
        message['Subject'] = 'Production report'
        message['From'] = mail_user_name
        message['To'] = ', '.join( list_of_mails ) 
        body_content = body
        message.attach(MIMEText(body_content, "html"))
        msg_body = message.as_string()
        count=0
        while True:
            try:
                server = SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(message['From'], app_password)
                server.sendmail(message['From'], list_of_mails, msg_body)
                server.quit()
                print("SM: Summary mail sent...")
            except Exception as e:
                print("SM: Unable to sent mail, ",e)
                time.sleep(3)
                count=count+1
                if count<3:
                    continue
                else:
                    break
            break

def trig_mail_content(shift):
    time.sleep(mail_trig_delay)
    table_list=[]
    required_row_list=[]
    required_column_list=[]
    no_data=True
    time_now=datetime.datetime.now()
    if shift=="C":
            dp=1
    else:
            dp=0
    time_now=time_now-datetime.timedelta(days=dp)
    date_today=time_now.strftime("%d-%m-%Y")
    summary_list=summary_dic[shift]
    table_list.append(summary_list)
    if os.path.isfile(filename_of_h_excel):
        wb=load_workbook(filename_of_h_excel)
        ws=wb[hourly_sheet_name]
        xl_headers=[]
        for i in ws[header_row_h]:
                xl_headers.append(i.value)
        date_today
        mr=ws.max_row
        for i in range(1,mr+1):
            if ws.cell(row=i,column=int(filedic["date_column"])).value==date_today:
                required_row_list.append(i)
        for i in summary_list:
            required_column_list.append(xl_headers.index(i)+1)

        raw_list=list(ws.rows)
        for r in required_row_list:
            no_data=False
            row_list=[]
            for c in required_column_list:
                val=raw_list[r-1][c-1].value
                if val is None:
                    val=0
                row_list.append(val)
            table_list.append(row_list)
        if no_data:
            print("TMC: As no data for the shift, mail not sent")
        else:
            send_mail(date_today,shift,table_list)
        wb.close()
    else:
        print(f"TMC: As no {filename_of_h_excel} found, mail not sent")


def schedule_thread():
    global unique_no_in_bytes
    try:
        for index_i,i in enumerate(hourly_list):
            schedule.every().day.at(i).do(hourly,str((index_i*2)+columns_before_pq_and_it_in_excel+1))
        for index_i,i in enumerate(end_timing_pos_list):
            end_timing_pos_list[index_i]=str((i*2)+columns_before_pq_and_it_in_excel+1)
        schedule.every().day.at(shiftA_start).do(trig_mail_content,"C")
        schedule.every().day.at(shiftB_start).do(trig_mail_content,"A")
        schedule.every().day.at(shiftC_start).do(trig_mail_content,"B")
        while True:            
            schedule.run_pending()
            time_now=datetime.datetime.now()
            current_time=datetime.time(time_now.hour,time_now.minute,time_now.second)
            if current_time<datetime.time(A[0],A[1],A[2]):
                dp=1
            else:
                dp=0
            time_now=time_now-datetime.timedelta(days=dp)
            date=time_now.strftime('%d%m')
            shift=get_shift(current_time)
            shift_no=ord(shift)-64
            unique_no=int(str(shift_no)+str(date))
            unique_no_in_bytes=unique_no.to_bytes(2,"big")
            time.sleep(scheduler_delay)

    except Exception as e:
        error_register_queue.put(f'MAT: {e}  {time.strftime("%d-%m-%Y_%I.%M.%S_%p")}\n')
        print(f'MAT: Error acquired and recorded. Error: {e}')
        time.sleep(error_wait)

for line_no in loaded_dict:
    line_thread_pro = threading.Thread(target=idle_time_calc,args=(line_no,),daemon=True)
    line_thread_pro.start()
    print(f'MT: Idle_time_calc thread started for line {line_data_dict[line_no]["l_n"]}')


socket_th = threading.Thread(target=socket_thread,daemon=True)
socket_th.start()
print("MT: Socket_thread thread started")
data_dec = threading.Thread(target=data_decode,daemon=True)
data_dec.start()
print("MT: Data_decode thread started")
move_ex_i = threading.Thread(target=move_excel_it,daemon=True)
move_ex_i.start()
print("MT: Move_excel_i thread started")
move_ex_h = threading.Thread(target=move_excel_h,daemon=True)
move_ex_h.start()
print("MT: Move_excel_h thread started")
google_ct_i = threading.Thread(target=google_chat_it,daemon=True)
google_ct_i.start()
print("MT: Google_chat_it thread started")
google_ct_h = threading.Thread(target=google_chat_h,daemon=True)
google_ct_h.start()
print("MT: Google_chat_h thread started")
schedule_th = threading.Thread(target=schedule_thread,daemon=True)
schedule_th.start()
print("MT: Schedule_thread thread started for hourly and mail trigger")
error_reg = threading.Thread(target=error_register,daemon=True)
error_reg.start()
print(f"MT: Error_register thread started")
dict_bac = threading.Thread(target=dict_backup,daemon=True)
dict_bac.start()
print(f"MT: Dict_backup thread started")

while True:
    time.sleep(60)